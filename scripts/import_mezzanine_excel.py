#!/usr/bin/env python3
"""Seed data/mezzanine.json and the body-parse cache from a hand-curated
Excel of CB/BW/EB issuances (used in the SME desk).

The Excel format has columns: 결의일, 납입일, 회사명, 종류, 회차,
발행금액, 표면이자율, 만기이자율, 만기일, ..., Put 행사일, Call 행사일,
Call 가능 비율, ...

Match strategy: (normalized 회사명, 종류, 회차) plus disambiguation by
납입일 when multiple candidates. Excel doesn't carry the DART rceptNo
or 종목코드, so name matching is the only join key.

Run from project root:
    python scripts/import_mezzanine_excel.py <xlsx>...
"""
import argparse
import json
import os
import re
import sys
from datetime import date, datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MEZZ_FILE = os.path.join(ROOT, "data", "mezzanine.json")
CACHE_FILE = os.path.join(ROOT, "data", "mezzanine-body-cache.json")


def _to_iso(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s or s == "-":
        return None
    m = re.match(r"^(\d{4})[\-./]?(\d{1,2})[\-./]?(\d{1,2})", s)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"


def _norm_name(s):
    if not s:
        return ""
    s = str(s).strip()
    # Strip 주식회사 prefix variants — Excel uses bare names, our JSON
    # comes from DART corp_name which often has them.
    s = re.sub(r"^\(주\)\s*|^주식회사\s*|^㈜\s*", "", s)
    s = re.sub(r"\s*\(주\)$|\s*주식회사$|\s*㈜$", "", s)
    return re.sub(r"\s+", "", s)


def _norm_type(s):
    if not s:
        return None
    s = str(s).strip().upper()
    if s.startswith("CB") or "전환사채" in s:
        return "CB"
    if s.startswith("BW") or "신주인수권" in s:
        return "BW"
    if s.startswith("EB") or "교환사채" in s:
        return "EB"
    return None


def _norm_round(v):
    if v is None or v == "-":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _pct(v):
    if v is None or v == "-":
        return None
    if isinstance(v, (int, float)):
        # Excel stores 0.5 for 50%, 0.3 for 30% etc.
        n = v * 100 if 0 < v <= 1 else v
        if n == int(n):
            return f"{int(n)}%"
        return f"{n:.1f}%"
    s = str(v).strip()
    if not s or s == "-":
        return None
    return s


def load_excel_rows(paths):
    """Yields normalized dicts from sheets named 사모_'YY년 / 공모_'YY년."""
    for path in paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sn in wb.sheetnames:
            if not (sn.startswith("사모_") or sn.startswith("공모_")):
                continue
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[2] is None:
                    continue
                btype = _norm_type(row[3])
                if not btype:
                    continue
                yield {
                    "sheet": sn,
                    "name": _norm_name(row[2]),
                    "raw_name": row[2],
                    "type": btype,
                    "round": _norm_round(row[4]),
                    "issueDate": _to_iso(row[1]),
                    "boardDate": _to_iso(row[0]),
                    "maturity": _to_iso(row[8]),
                    "putStart": _to_iso(row[11]),
                    "callStart": _to_iso(row[13]),
                    "callPct": _pct(row[15]),
                }


def build_index(mezz):
    """Three indexes for progressive matching:

    1. (norm_name, type, round) — strict, used first.
    2. (norm_name) — name-only fallback when round mismatches.
    3. (issueDate, type, round) — name-INDEPENDENT fallback. Catches
       companies whose ticker-file name is stale (사명 변경 미반영):
       e.g. ticker 377030 still reads "비트맥스" but Excel has its
       current name "맥스트". The CB itself was filed on the same
       issueDate with the same type and round, so the date+type+round
       triple matches uniquely.
    """
    by_full = {}
    by_name = {}
    by_date = {}
    for ticker, entry in mezz.get("mezzanine", {}).items():
        name_n = _norm_name(entry.get("name"))
        for iss in entry.get("issuances", []):
            rnd = _norm_round(iss.get("round"))
            issd = iss.get("issueDate")
            full_key = (name_n, iss["type"], rnd)
            by_full.setdefault(full_key, []).append((ticker, iss))
            by_name.setdefault(name_n, []).append((ticker, iss))
            if issd and rnd:
                by_date.setdefault((issd, iss["type"], rnd), []).append(
                    (ticker, iss)
                )
    return by_full, by_name, by_date


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", nargs="+")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    with open(MEZZ_FILE, encoding="utf-8") as f:
        mezz = json.load(f)
    try:
        with open(CACHE_FILE, encoding="utf-8") as f:
            cache = json.load(f)
    except FileNotFoundError:
        cache = {}

    by_full, by_name, by_date = build_index(mezz)
    rows = list(load_excel_rows(args.excel))
    print(f"Loaded {len(rows):,} Excel rows", flush=True)

    matched = 0
    matched_by = {"full": 0, "name_only": 0, "date_fallback": 0}
    unmatched_examples = []
    unmatched_count = 0
    by_unmatch_reason = {"name_missing": 0, "round_mismatch": 0}

    for row in rows:
        candidates = by_full.get((row["name"], row["type"], row["round"]), [])
        match_kind = "full"
        if not candidates:
            # Fallback 1: name + type, ignore round.
            pool = [
                (t, i) for (t, i) in by_name.get(row["name"], [])
                if i["type"] == row["type"]
            ]
            if pool and row["issueDate"]:
                for cand in pool:
                    if cand[1].get("issueDate") == row["issueDate"]:
                        candidates = [cand]
                        match_kind = "name_only"
                        break
        if not candidates and row["issueDate"] and row["round"]:
            # Fallback 2: name-INDEPENDENT match on (issueDate, type,
            # round). Catches stale ticker-file names (사명 변경).
            candidates = list(by_date.get(
                (row["issueDate"], row["type"], row["round"]), []
            ))
            if candidates:
                match_kind = "date_fallback"
        if not candidates:
            unmatched_count += 1
            if by_name.get(row["name"]):
                by_unmatch_reason["round_mismatch"] += 1
            else:
                by_unmatch_reason["name_missing"] += 1
                if len(unmatched_examples) < 8:
                    unmatched_examples.append(row)
            continue

        # Disambiguate by issueDate if multiple round-matches.
        chosen = None
        if len(candidates) > 1 and row["issueDate"]:
            for cand in candidates:
                if cand[1].get("issueDate") == row["issueDate"]:
                    chosen = cand
                    break
        if not chosen:
            chosen = candidates[0]

        ticker, iss = chosen
        matched_by[match_kind] += 1
        # Apply put/call to issuance.
        if row["putStart"]:
            iss["putStart"] = row["putStart"]
            iss["putSource"] = "excel"
        if row["callStart"]:
            iss["callStart"] = row["callStart"]
        if row["callPct"]:
            iss["callPct"] = row["callPct"]

        # Also write to body cache keyed by rceptNo so a future workflow
        # run doesn't overwrite the seeded values.
        rcept_no = iss.get("rceptNo")
        if rcept_no:
            entry = dict(cache.get(rcept_no, {}))
            entry["outcome"] = entry.get("outcome", "ok")
            if row["putStart"]:
                entry["putStart"] = row["putStart"]
            if row["callStart"]:
                entry["callStart"] = row["callStart"]
            if row["callPct"]:
                entry["callPct"] = row["callPct"]
            entry["source"] = "excel"
            cache[rcept_no] = entry

        matched += 1

    print(
        f"\nMatched: {matched:,} / Unmatched: {unmatched_count:,} "
        f"({100.0 * matched / len(rows):.1f}% match rate)",
        flush=True,
    )
    print(f"  matched by:      {matched_by}", flush=True)
    print(f"  unmatch reasons: {by_unmatch_reason}", flush=True)
    if unmatched_examples:
        print("\n  Sample unmatched (likely missing from DART scan):")
        for r in unmatched_examples:
            print(
                f"    {r['issueDate']}  {r['raw_name']}  {r['type']} 제{r['round']}회"
            )

    if args.dry_run:
        print("\n(dry-run: no files written)")
        return

    with open(MEZZ_FILE, "w", encoding="utf-8") as f:
        json.dump(mezz, f, ensure_ascii=False, separators=(",", ":"))
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, separators=(",", ":"),
                  sort_keys=True)
    print(f"\nWrote {MEZZ_FILE}")
    print(f"Wrote {CACHE_FILE}  ({len(cache):,} cached entries)")


if __name__ == "__main__":
    main()
